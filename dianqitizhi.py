import os
import re
import win32com.client as win32

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, NamedStyle

"""
将厂家的excel的返资格式转换为给电气提资的格式
"""

# 设备编号,型号列
EXTRACT_ITEMS = {'设备编号': ['设备编号'], '设备名称': ['设备名称'], '设备型号': ['设备型号']}
# 设备参数列
EXTRACT_VALUES = {'功率': ['功率', '功耗','用电量'], '电源': ['电源', '电压','用电规格','用电']}
# 厂家反馈的参数列的关键词,
CHOOSE_VALUES_TOTAL_SIGN_WORD = {'参数': ['投标', '反馈', '选型','招标']}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def init_df(df):
    """
    初始化体质sheet中的df
    :param df:
    :return:
    """
    if not is_df_valid(df):
        raise TypeError('Input df is not a valid pd.DataFrame.')

    def drop_valid_str(string):
        # 替换字符串中的括号以及删除空格
        if not isinstance(string, str):
            return string
        else:
            newString = re.sub('[（(].*[）)]', '', string)
            return newString.strip()

    def init_sign_words(string):
        """
        规范一些标致性的词
        :param string:
        :return:
        """
        if not isinstance(string, str):
            return string
        for itmsDict in [EXTRACT_ITEMS, EXTRACT_VALUES, CHOOSE_VALUES_TOTAL_SIGN_WORD]:
            for signWord, replaceWordList in itmsDict.items():
                for replaceWord in replaceWordList:
                    if replaceWord in string:
                        return signWord
        return string

    # 首先替换字符串中的括号以及空格
    df = df.applymap(lambda x: drop_valid_str(x) if isinstance(x, str) else x)
    # 替换关键词
    df = df.applymap(lambda x: init_sign_words(x) if isinstance(x, str) else x)
    return df


def is_df_valid(df):
    """
    检测df
    :param df:
    :return:
    """
    if df is None:
        return False
    if not isinstance(df, pd.DataFrame):
        return False
    if len(df.index) == 0:
        return False
    return True


def check_contain_chinese(check_str):
    for ch in check_str:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


def process_changjia_df(df):
    """
    处理厂家的df
    :param df:
    :return:
    """
    if not is_df_valid(df):
        return

    # 初始化处理厂家文件内容:1.删除空白值,填充空值
    processedDf = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    processedDf.fillna(value='', inplace=True)
    # 提取内容
    # todo 后面需要加入类别转换判断
    for item, _ in EXTRACT_ITEMS.items():
        # 找到首次出现关键项的行
        showRow = None
        for row, rowSeries in processedDf.iterrows():
            if item in rowSeries.values:
                showRow = row
                break
        if showRow is not None:
            break
    # 选择特定的列
    extractItemSignWord = set(EXTRACT_ITEMS.keys())
    extractValuesSignWord = set(EXTRACT_VALUES.keys())
    chooseValuesTotalSignWord = '参数'
    chooseCol = pd.DataFrame()

    for col, colSeries in processedDf.iteritems():
        if extractItemSignWord & set(colSeries.values) != set():
            colName = list(extractItemSignWord & set(colSeries.values))[0]
            extentColDf = pd.DataFrame([colSeries]).T
            extentColDf.columns = [colName]
            chooseCol = pd.concat([chooseCol, extentColDf], axis=1)
            continue

        if (chooseValuesTotalSignWord in colSeries.values) and (extractValuesSignWord & set(colSeries.values) != set()):
            colName = list(extractValuesSignWord & set(colSeries.values))[0]
            extentColDf = pd.DataFrame([colSeries]).T
            extentColDf.columns = [colName]
            chooseCol = pd.concat([chooseCol, extentColDf], axis=1)
            continue

    if chooseCol.empty:
        # todo 没有可以选择的行,如何进行提示
        return pd.DataFrame()
    # 筛选设备编号:设备编号列不含中文
    chooseCol = chooseCol[
        chooseCol['设备编号'].apply(lambda x: not check_contain_chinese(x) if isinstance(x, str) else False)]
    chooseCol  = chooseCol[chooseCol['设备编号'].apply(lambda x:True if len(x)>1 else False)]
    chooseCol.drop_duplicates(['设备编号'], keep='first', inplace=True)
    return chooseCol


def trans_xls_2_xlsx(xls_path):
    '''
    将xls转换为xlsx
    :param xls_path:
    :return:
    '''
    if not os.path.isfile(xls_path):
        raise 1
    fileType = os.path.splitext(xls_path)[1]
    if '.xls' != fileType.lower():
        raise 1
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(xls_path)
    # xlsx: FileFormat=51
    # xls:  FileFormat=56,
    # 后缀名的大小写不通配，需按实际修改：xls，或XLS
    dirName = os.path.basename(xls_path)
    saveFolder = os.path.join(os.path.dirname(xls_path), 'temp')
    if not os.path.exists(saveFolder):
        os.mkdir(saveFolder)
    savePath = os.path.join(saveFolder,
                            os.path.splitext(dirName)[0] + os.path.splitext(dirName)[1].lower().replace('xls', 'xlsx'))
    if os.path.exists(savePath):
        try:
            os.remove(savePath)
        except:
            raise OSError('f{savePath}被打开了,请关闭后重新运行程序.')
    wb.SaveAs(savePath, FileFormat=51)  # 我这里原文件是大写
    wb.Close()
    excel.Application.Quit()
    return savePath


def load_excel(path):
    """
    加载excel文件
    :param path:
    :return: list of pd.DataFrame
    """
    if not os.path.exists(path):
        # todo 这里可以弹出提示框
        return []
    xls = pd.ExcelFile(path)
    sheetNames = xls.sheet_names
    extraCol = 'Macro1'
    if extraCol in sheetNames:
        sheetNames.remove(extraCol)
    if len(sheetNames) == 0:
        raise ValueError(f"文件夹{path}为空工作簿.")
    elif len(sheetNames) == 1:
        return [read_merged_excel(path, sheetname=sheetNames[0])]
    else:
        return [read_merged_excel(path, sheetname=sheet) for sheet in sheetNames]


def read_merged_excel(path, sheetname, is_first_row_column=True):
    """
    读取合并的excel文件中的一个文件表
    :param path:
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name(sheetname)
    df = pd.DataFrame(wb[sheetname].values)
    for crange in sheet.merged_cells.ranges:
        clo, rlo, chi, rhi = crange.bounds
        for row in range(rlo, rhi + 1):
            for col in range(clo, chi + 1):
                cellValue = sheet.cell(row, col).value
                if cellValue != None:
                    value = cellValue
                    break
            if value != None:
                break
        df.loc[rlo - 1:rhi - 1, clo - 1:chi - 1] = value
    if is_first_row_column:
        df.columns = df.loc[0, :]
        df = df.drop(index=[0])
    else:
        pass
    df.reset_index(drop=True, inplace=True)
    return df


def combine_system(biddingSystemDf, exportPath):
    """
    将提资整理之后的数据按照系统分组,是按照系统设备编号来分组。
        多联机系统:以F或是V开头
        A系统:设备的序号编号以A开头
        B系统:设备的序号编号以B开头
        其他系统:不满足以上的条件
    :param biddingSystemDf:
    :return:
    """
    if not is_df_valid(biddingSystemDf):
        return
    EquipmentColumnName = '设备编号'
    systemColumns = biddingSystemDf.columns
    if EquipmentColumnName not in systemColumns:
        return
    # 多联机系统设备
    combinedEquipmentSystemDf = biddingSystemDf[(biddingSystemDf[EquipmentColumnName].str.startswith('V', na=False)) | (
        biddingSystemDf[EquipmentColumnName].str.startswith('F', na=False))]
    restbiddingSystemDf = biddingSystemDf[~((biddingSystemDf[EquipmentColumnName].str.startswith('V', na=False)) | (
        biddingSystemDf[EquipmentColumnName].str.startswith('F', na=False)))]

    # 空调A端设备
    pointAEquipmentSystemDf = restbiddingSystemDf[restbiddingSystemDf[EquipmentColumnName].str.contains('-A')]
    restbiddingSystemDf = restbiddingSystemDf[~(restbiddingSystemDf[EquipmentColumnName].str.contains('-A'))]

    # 空调B端设备
    pointBEquipmentSystemDf = restbiddingSystemDf[restbiddingSystemDf[EquipmentColumnName].str.contains('-B')]

    # 其他设备
    restbiddingSystemDf = restbiddingSystemDf[~(restbiddingSystemDf[EquipmentColumnName].str.contains('-B'))]

    # 合并输出
    xlsxWorkBook = openpyxl.workbook.Workbook()
    workSheet = xlsxWorkBook.create_sheet('返给电气')

    # 然后如下设置：
    # 设置表头字体居中
    # 创建单元格样式对象
    headerCellStyle = NamedStyle(name='headerCellStyle')
    headerCellStyle.alignment = Alignment(horizontal='center', vertical='center')  # 水平垂直居

    nowRow = 1
    # 写列名
    for colNum in range(len(systemColumns)):
        workSheet.cell(nowRow, colNum + 1).value = systemColumns[colNum]
        workSheet.cell(nowRow, colNum + 1).style = headerCellStyle
    nowRow += 1
    startColumnsName = 'A'
    endColumnsName = get_column_letter(len(systemColumns))

    # 写A系统
    if len(pointAEquipmentSystemDf.index) != 0:
        pointAEquipmentSystemDfSize = pointAEquipmentSystemDf.shape

        # 写A系统标题
        workSheet.merge_cells(f'{startColumnsName}{nowRow}:{endColumnsName}{nowRow}')
        workSheet.cell(nowRow, 1).value = '空调A端设备'
        workSheet.cell(nowRow, 1).style = headerCellStyle
        nowRow += 1

        for row in range(pointAEquipmentSystemDfSize[0]):
            for colNum in range(pointAEquipmentSystemDfSize[1]):
                workSheet.cell(nowRow + row, colNum + 1).value = pointAEquipmentSystemDf.iloc[row, colNum]

        nowRow += pointAEquipmentSystemDfSize[0]

    # 写B系统

    if len(pointBEquipmentSystemDf.index) != 0:
        pointBEquipmentSystemDfSize = pointBEquipmentSystemDf.shape

        # 写B系统标题
        workSheet.merge_cells(f'{startColumnsName}{nowRow}:{endColumnsName}{nowRow}')
        workSheet.cell(nowRow, 1).value = '空调B端设备'
        workSheet.cell(nowRow, 1).style = headerCellStyle
        nowRow += 1

        for row in range(pointBEquipmentSystemDfSize[0]):
            for colNum in range(pointBEquipmentSystemDfSize[1]):
                workSheet.cell(nowRow + row, colNum + 1).value = pointBEquipmentSystemDf.iloc[row, colNum]

        nowRow += pointBEquipmentSystemDfSize[0]

    # 多联机系统设备
    if len(combinedEquipmentSystemDf.index) != 0:
        combinedEquipmentSystemDfSize = combinedEquipmentSystemDf.shape

        # 多联机系统设备标题
        workSheet.merge_cells(f'{startColumnsName}{nowRow}:{endColumnsName}{nowRow}')
        workSheet.cell(nowRow, 1).value = '多联机系统设备标题'
        workSheet.cell(nowRow, 1).style = headerCellStyle
        nowRow += 1

        for row in range(combinedEquipmentSystemDfSize[0]):
            for colNum in range(combinedEquipmentSystemDfSize[1]):
                workSheet.cell(nowRow + row, colNum + 1).value = combinedEquipmentSystemDf.iloc[row, colNum]

        nowRow += combinedEquipmentSystemDfSize[0]

    # 其他系统

    if len(restbiddingSystemDf.index) != 0:
        restbiddingSystemDfSize = restbiddingSystemDf.shape

        # 其他系统
        workSheet.merge_cells(f'{startColumnsName}{nowRow}:{endColumnsName}{nowRow}')
        workSheet.cell(nowRow, 1).value = '其他系统'
        workSheet.cell(nowRow, 1).style = headerCellStyle
        nowRow += 1

        for row in range(restbiddingSystemDfSize[0]):
            for colNum in range(restbiddingSystemDfSize[1]):
                workSheet.cell(nowRow + row, colNum + 1).value = restbiddingSystemDf.iloc[row, colNum]

    xlsxWorkBook.save(exportPath)


def combine_equipment(equipmentDf):
    """
    合并设备编号:
    当设备名称相同,设备编号前面的大型相同,功率相同时,将这些设备合并.
    :param equipmentDf:
    :return:
    """

    # 属性字段
    equipmentName = '设备名称'
    equipmentPower = '功率'
    equipmentPowerSupply = '电源'
    # 计算字段
    equipmentNum = '设备编号'
    equipmentQuantity = '数量'
    equipmentPowerSum = '合计功率'
    # 判断是否有效
    if not is_df_valid(equipmentDf):
        return
    equipmentDf.reset_index(drop=True, inplace=True)

    # 如果数量列不存在,则添加数量列
    if equipmentQuantity not in equipmentDf.columns:
        equipmentDf[equipmentQuantity] = 1
    else:
        equipmentDf[equipmentQuantity].fillna(value=1,inplace=True)


    # 如果设备名称列不存在,则设备名称列
    if equipmentName not in equipmentDf.columns:
        equipmentDf[equipmentName] = ''
    else:
        equipmentDf[equipmentName].fillna(value='',inplace=True)


    # 如果合计功率不存在,则添加合计功率
    if equipmentPowerSum not in equipmentDf.columns:
        equipmentDf[equipmentPowerSum] = equipmentDf[equipmentPower]


    # 设备编号的辅助列来判断是不是一个设备型号和端
    equipmentNumSign = '设备编号辅助'
    equipmentDf[equipmentNumSign] = [None if not isinstance(num, str) else
                                     None if '-' not in num else num[0:num.find('-') + 2]
                                     for num in equipmentDf[equipmentNum]]
    equipmentNumNum = '设备编号标号'
    equipmentDf[equipmentNumNum] = [None if not isinstance(num, str) else
                                    None if '-' not in num else num[num.find('-') + 2:].replace('、',',')
                                    for num in equipmentDf[equipmentNum]]
    equipmentColumns = [col.strip() for col in equipmentDf.columns]
    equipmentSignDict = {equipmentName: '', equipmentNumSign: '', equipmentPower: '', equipmentPowerSupply: ''}
    equipmentValueDict = {equipmentNumNum: '', equipmentQuantity: '', equipmentPowerSum: ''}
    for signCol, value in equipmentSignDict.items():
        for index, colName in enumerate(equipmentColumns):
            if signCol == colName:
                equipmentSignDict[signCol] = index
                break
    for valueCol, value in equipmentValueDict.items():
        for index, colName in enumerate(equipmentColumns):
            if valueCol == colName:
                equipmentValueDict[valueCol] = index
                break

    # 处理前后行特殊字段数值相同,将计算字段相加


    # 将编号打开
    def openNum(numStr):
        if not isinstance(numStr, str):
            return []
        if numStr == '':
            return []
        aList = []
        for i in [range(int(i.split('~')[0]), int(i.split('~')[-1]) + 1) for i in numStr.split(',')]:
            for j in i:
                aList.append(j)
        return aList

    # 将编号关闭
    def closeNum(numList):
        if not isinstance(numList, list):
            return ''
        if len(numList) == 0:
            return ''
        sortA = list(sorted(numList))
        ListA = []
        while 1:
            nowPart = []
            for index, valueA in enumerate(sortA):
                if (index != len(sortA) - 1) and (valueA + 1 != sortA[index + 1]):
                    nowPart.append(valueA)
                    break
                nowPart.append(valueA)
            if len(nowPart) == 1:
                nowPartStr = f'{nowPart[0]}'
            else:
                nowPartStr = f'{nowPart[0]}~{nowPart[-1]}'
            ListA.append(nowPartStr)
            sortA = sortA[index + 1:]
            if len(sortA) == 0:
                break
        return ','.join(ListA)

    equipmentDfIndex = list(equipmentDf.index)
    dropIndex = []
    for index in equipmentDfIndex:
        # 重新计算设备编号和设备标识编号
        equipmentDf[equipmentNumNum] = [None if not isinstance(num, str) else
                                        None if '-' not in num else num[num.find('-') + 2:].replace('、',',')
                                        for num in equipmentDf[equipmentNum]]

        equipmentDf[equipmentNumSign] = [None if not isinstance(num, str) else
                                         None if '-' not in num else num[0:num.find('-') + 2]
                                         for num in equipmentDf[equipmentNum]]

        # 设备编号的合并
        firstNumNum = equipmentDf.iloc[index, equipmentValueDict[equipmentNumNum]]
        firstNumNumRange = openNum(firstNumNum)
        equipmentDf.loc[index, equipmentQuantity] = len(firstNumNumRange)
        equipmentDf.loc[index,equipmentPowerSum] = len(firstNumNumRange) * equipmentDf.loc[index,equipmentPower]

        # 最后一行
        if index == equipmentDf.shape[0] - 1:
            continue

        secondNumNum = equipmentDf.iloc[index + 1, equipmentValueDict[equipmentNumNum]]
        secondNumNumRange = openNum(secondNumNum)
        allNumNumRange = firstNumNumRange
        allNumNumRange.extend(secondNumNumRange)
        allNumNumStr = closeNum(allNumNumRange)
        # 修改索引列的数量和合计功率

        for signColIndex in equipmentSignDict.keys():
            sameRowFlag = False
            if equipmentDf.loc[index, signColIndex] in [None, '']:
                break
            if equipmentDf.loc[index, signColIndex] != equipmentDf.loc[index + 1, signColIndex]:
                break
            sameRowFlag = True
        if not sameRowFlag:
            continue

        equipmentNumCombine = equipmentDf.iloc[index, equipmentSignDict[equipmentNumSign]] + allNumNumStr

        # 设备数量求和
        equipmengQuantityCombine = len(allNumNumRange)

        # 合并功率
        equipmentPowerSumCombine = float(equipmentDf.loc[index, equipmentPower])*equipmengQuantityCombine

        equipmentDf.loc[index + 1, equipmentNum] = equipmentNumCombine
        equipmentDf.loc[index + 1, equipmentQuantity] = equipmengQuantityCombine
        equipmentDf.loc[index + 1, equipmentPowerSum] = equipmentPowerSumCombine
        # 删掉的列
        dropIndex.append(index)
    # 删除行
    equipmentDf = equipmentDf.drop(index=dropIndex)

    # 删除负责列
    del equipmentDf[equipmentNumSign]
    del equipmentDf[equipmentNumNum]
    return equipmentDf


def modify_path(file_path):
    """
    标准化路径
    :param file_path:
    :return:
    """
    if not os.path.isfile(file_path):
        raise FileExistsError(f'{file_path}不存在.')
    fileType = os.path.splitext(file_path)[1]
    if fileType.lower() not in ['.xls', '.xlsx']:
        raise TypeError(f'{file_path}type is {fileType} NOT one of excel type.')
    if fileType.lower() == '.xlsx':
        return file_path
    else:
        return trans_xls_2_xlsx(file_path)


if __name__ == '__main__':
    dataBasePath = r"data/提电气资料（环控、动力、FAS、BAS/厂家返/5223--都市工业园站水阀选型--市政院-19.8.2.xls"
    dataPath = os.path.join(BASE_DIR,dataBasePath)
    newPth = modify_path(dataPath)
    # # read_merged_excel(path1,sheetName)
    sheetListDf = load_excel(newPth)
    if sheetListDf is not None:
        if len(sheetListDf) == 1:
            sheetDf = init_df(sheetListDf[0])
            modifyChangjiaDf = process_changjia_df(sheetDf)
        else:
            modifyChangjiaDf = pd.DataFrame()
            for sheetDf in sheetListDf:
                sheetDf = init_df(sheetDf)
                oneChangjiaDf = process_changjia_df(sheetDf)
                modifyChangjiaDf = pd.concat([modifyChangjiaDf, oneChangjiaDf])
    modifyChangjiaDf = combine_equipment(modifyChangjiaDf)
    path2BasePath = r"data/提电气资料（环控、动力、FAS、BAS/厂家返/523都市工业园站---------多联空调参数表20190707（再次确认参数）23.xls"
    path2 = os.path.join(BASE_DIR,path2BasePath)
    combine_system(modifyChangjiaDf,path2)
